from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import pandas as pd
import pyreadstat
import requests
from openpyxl import load_workbook

API = "https://api.osf.io/v2"
NODE_ID = "ha4um"
VIEW_ONLY = "54e9003e0690429f82f1c708545dc1c5"
MAX_FILE_BYTES = 250 * 1024 * 1024
MAX_TOTAL_BYTES = 800 * 1024 * 1024
SUPPORTED = {".csv", ".tsv", ".xlsx", ".xls", ".sav", ".dta", ".r", ".rmd", ".txt", ".json", ".zip", ".docx", ".pdf"}
SENSITIVE = re.compile(r"(^|_)(email|name|phone|address|ip|workerid|mturkid|prolificid|participantid)($|_)", re.I)
BEHAVIOR_HINT = re.compile(r"(choice|advisor|algorithm|human|responsib|trust|disclaimer|condition|domain|medical|financial|participant|subject|response)", re.I)


def add_view_only(url: str) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["view_only"] = VIEW_ONLY
    return urlunparse(parsed._replace(query=urlencode(query)))


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def get_json(session: requests.Session, url: str) -> dict:
    response = session.get(add_view_only(url), timeout=(30, 180))
    response.raise_for_status()
    value = response.json()
    if not isinstance(value, dict):
        raise RuntimeError("expected_json_object")
    return value


def paginate(session: requests.Session, url: str) -> list[dict]:
    output: list[dict] = []
    next_url: str | None = url
    seen: set[str] = set()
    while next_url:
        target = add_view_only(next_url)
        if target in seen:
            raise RuntimeError("repeated_pagination_url")
        seen.add(target)
        response = session.get(target, timeout=(30, 180))
        response.raise_for_status()
        value = response.json()
        output.extend(item for item in value.get("data", []) if isinstance(item, dict))
        next_url = (value.get("links") or {}).get("next")
    return output


def flatten_files(session: requests.Session, url: str) -> list[dict]:
    queue = [url]
    output: list[dict] = []
    seen: set[str] = set()
    while queue:
        current = queue.pop(0)
        key = add_view_only(current)
        if key in seen:
            continue
        seen.add(key)
        for item in paginate(session, current):
            attrs = item.get("attributes") or {}
            links = item.get("links") or {}
            relationships = item.get("relationships") or {}
            kind = attrs.get("kind")
            files_url = ((((relationships.get("files") or {}).get("links") or {}).get("related") or {}).get("href"))
            row = {
                "id": item.get("id"),
                "name": attrs.get("name"),
                "kind": kind,
                "size": attrs.get("size"),
                "materialized_path": attrs.get("materialized_path"),
                "date_modified": attrs.get("date_modified"),
                "download_url": links.get("download"),
                "files_url": files_url,
            }
            if kind == "folder" and files_url:
                queue.append(files_url)
            elif kind == "file":
                output.append(row)
    return output


def download(session: requests.Session, url: str, path: Path) -> dict:
    path.parent.mkdir(parents=True, exist_ok=True)
    total = 0
    with session.get(add_view_only(url), timeout=(30, 300), stream=True, allow_redirects=True) as response:
        response.raise_for_status()
        with path.open("wb") as fh:
            for chunk in response.iter_content(1024 * 1024):
                if not chunk:
                    continue
                total += len(chunk)
                if total > MAX_FILE_BYTES:
                    raise RuntimeError("file_exceeds_limit")
                fh.write(chunk)
        return {"final_url": response.url, "content_type": response.headers.get("Content-Type"), "size_bytes": total, "sha256": sha256(path)}


def schema_from_dataframe(frame: pd.DataFrame) -> dict:
    columns = [str(col) for col in frame.columns]
    sensitive = [col for col in columns if SENSITIVE.search(col)]
    behavioral = [col for col in columns if BEHAVIOR_HINT.search(col)]
    return {
        "rows": int(len(frame)),
        "columns": len(columns),
        "column_names": columns,
        "sensitive_column_candidates": sensitive,
        "behavioral_column_candidates": behavioral,
        "participant_level_candidate": bool(behavioral and len(frame) >= 50),
    }


def inspect_file(path: Path) -> dict:
    suffix = path.suffix.lower()
    result = {"signature_hex": path.read_bytes()[:16].hex(), "suffix": suffix}
    if suffix == ".csv":
        frame = pd.read_csv(path, nrows=None, low_memory=False)
        result.update(schema_from_dataframe(frame))
    elif suffix == ".tsv":
        frame = pd.read_csv(path, sep="\t", nrows=None, low_memory=False)
        result.update(schema_from_dataframe(frame))
    elif suffix in {".xlsx", ".xls"}:
        if suffix == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=False)
            sheets = []
            for sheet in workbook.worksheets:
                iterator = sheet.iter_rows(values_only=True)
                try:
                    header = next(iterator)
                except StopIteration:
                    header = tuple()
                rows = sum(1 for _ in iterator)
                columns = ["" if value is None else str(value) for value in header]
                sheets.append({"sheet": sheet.title, "rows": rows, "columns": len(columns), "column_names": columns, "sensitive_column_candidates": [c for c in columns if SENSITIVE.search(c)], "behavioral_column_candidates": [c for c in columns if BEHAVIOR_HINT.search(c)]})
            workbook.close()
            result["sheets"] = sheets
        else:
            result["structural_parse"] = "xls_not_parsed_without_legacy_engine"
    elif suffix == ".sav":
        frame, meta = pyreadstat.read_sav(path)
        result.update(schema_from_dataframe(frame))
        result["column_labels"] = dict(zip(meta.column_names, meta.column_labels or []))
    elif suffix == ".dta":
        frame, meta = pyreadstat.read_dta(path)
        result.update(schema_from_dataframe(frame))
        result["column_labels"] = dict(zip(meta.column_names, meta.column_labels or []))
    elif suffix in {".r", ".rmd"}:
        text = path.read_text(encoding="utf-8", errors="replace")
        result.update({"line_count": len(text.splitlines()), "code_executed": False, "data_file_references": sorted(set(re.findall(r"[A-Za-z0-9_./ -]+\.(?:csv|sav|dta|xlsx|rds|RData)", text, re.I)))})
    else:
        result["structural_parse"] = "not_executed_or_not_required"
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)
    files_dir = args.output / "files"
    session = requests.Session()
    session.headers.update({"User-Agent": "ExplainabilityBiasOpenEvidence/2.1 OSF-lawful-data-audit", "Accept": "application/vnd.api+json, application/json, */*"})

    node = get_json(session, f"{API}/nodes/{NODE_ID}/")
    node_data = node.get("data") or {}
    node_attrs = node_data.get("attributes") or {}
    providers_url = (((((node_data.get("relationships") or {}).get("files") or {}).get("links") or {}).get("related") or {}).get("href"))
    if not providers_url:
        raise RuntimeError("files_relationship_missing")
    providers = paginate(session, providers_url)
    files: list[dict] = []
    for provider in providers:
        relationships = provider.get("relationships") or {}
        files_url = ((((relationships.get("files") or {}).get("links") or {}).get("related") or {}).get("href"))
        if files_url:
            files.extend(flatten_files(session, files_url))

    downloaded = []
    errors = []
    total = 0
    for index, row in enumerate(files, 1):
        name = str(row.get("name") or row.get("id"))
        suffix = Path(name).suffix.lower()
        size = row.get("size") if isinstance(row.get("size"), int) else 0
        if suffix not in SUPPORTED or not row.get("download_url"):
            continue
        if size > MAX_FILE_BYTES or total + size > MAX_TOTAL_BYTES:
            errors.append({"name": name, "error": "download_budget_or_file_limit"})
            continue
        safe = re.sub(r"[^A-Za-z0-9._() -]+", "_", Path(name).name)
        path = files_dir / f"{index:04d}_{safe}"
        try:
            transfer = download(session, row["download_url"], path)
            audit = inspect_file(path)
        except Exception as exc:
            errors.append({"name": name, "error": f"{type(exc).__name__}:{exc}"})
            continue
        total += transfer["size_bytes"]
        downloaded.append({"osf_file": row, "local_file": str(path.relative_to(args.output)), "transfer": transfer, "audit": audit})

    participant_candidates = [item for item in downloaded if item["audit"].get("participant_level_candidate") or any(sheet.get("behavioral_column_candidates") for sheet in item["audit"].get("sheets", []))]
    sensitive_candidates = [item for item in downloaded if item["audit"].get("sensitive_column_candidates") or any(sheet.get("sensitive_column_candidates") for sheet in item["audit"].get("sheets", []))]
    report = {
        "official_api": API,
        "node_id": NODE_ID,
        "view_only_used": True,
        "node_title": node_attrs.get("title"),
        "node_public": node_attrs.get("public"),
        "providers": len(providers),
        "files_discovered": len(files),
        "files_downloaded": len(downloaded),
        "downloaded_bytes": total,
        "participant_level_candidate_files": [item["osf_file"].get("name") for item in participant_candidates],
        "sensitive_column_candidate_files": [item["osf_file"].get("name") for item in sensitive_candidates],
        "errors": errors,
        "downloaded": downloaded,
        "code_executed": False,
        "completed": True,
    }
    (args.output / "audit_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    sums = []
    for file_path in sorted(path for path in args.output.rglob("*") if path.is_file() and path.name != "sha256sums.txt"):
        sums.append(f"{sha256(file_path)}  {file_path.relative_to(args.output)}")
    (args.output / "sha256sums.txt").write_text("\n".join(sums) + "\n", encoding="utf-8")
    print(json.dumps({k: report[k] for k in ("node_title", "files_discovered", "files_downloaded", "participant_level_candidate_files", "sensitive_column_candidate_files", "errors")}, ensure_ascii=False, indent=2))
    if not participant_candidates:
        raise SystemExit("No participant-level candidate file was verified")


if __name__ == "__main__":
    main()
