from __future__ import annotations

import argparse
import hashlib
import json
import mimetypes
import os
import re
from pathlib import Path
from urllib.parse import urlparse

import requests

MAX_FILE_BYTES = 250 * 1024 * 1024
MAX_TOTAL_BYTES = 500 * 1024 * 1024
ALLOWED_HOST_SUFFIXES = (
    "zenodo.org",
    "api.osf.io",
    "osf.io",
    "storage.googleapis.com",
    "amazonaws.com",
    "cloudfront.net",
)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def allowed_host(url: str) -> bool:
    host = (urlparse(url).hostname or "").lower()
    return any(host == s or host.endswith("." + s) for s in ALLOWED_HOST_SUFFIXES)


def checked_get(session: requests.Session, url: str, *, stream: bool = False) -> requests.Response:
    if not url.startswith("https://"):
        raise RuntimeError(f"non_https_url:{url}")
    if not allowed_host(url):
        raise RuntimeError(f"host_not_allowed:{urlparse(url).hostname}")
    r = session.get(url, timeout=(30, 240), allow_redirects=True, stream=stream)
    if not allowed_host(r.url):
        raise RuntimeError(f"redirect_host_not_allowed:{urlparse(r.url).hostname}")
    r.raise_for_status()
    return r


def download(session: requests.Session, url: str, path: Path) -> dict:
    r = checked_get(session, url, stream=True)
    total = 0
    with path.open("wb") as f:
        for chunk in r.iter_content(1024 * 1024):
            if not chunk:
                continue
            total += len(chunk)
            if total > MAX_FILE_BYTES:
                raise RuntimeError("file_too_large")
            f.write(chunk)
    if total == 0:
        raise RuntimeError("empty_file")
    return {
        "requested_url": url,
        "final_url": r.url,
        "status_code": r.status_code,
        "content_type": (r.headers.get("content-type") or "").lower(),
        "size_bytes": total,
        "sha256": sha256_file(path),
    }


def inspect_file(path: Path) -> dict:
    suffix = path.suffix.lower()
    head = path.read_bytes()[:16]
    info = {"extension": suffix, "magic_hex": head.hex(), "structural_status": "not_parsed"}
    try:
        if suffix in {".csv", ".tsv"}:
            import csv
            dialect = "excel-tab" if suffix == ".tsv" else "excel"
            with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
                reader = csv.reader(f, dialect=dialect)
                rows = []
                for i, row in enumerate(reader):
                    rows.append(row)
                    if i >= 10000:
                        break
            info.update({
                "structural_status": "parsed_delimited",
                "sampled_rows": max(0, len(rows) - 1),
                "columns": rows[0] if rows else [],
                "sample_truncated": len(rows) >= 10001,
            })
        elif suffix in {".xlsx", ".xlsm"}:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
            sheets = []
            for ws in wb.worksheets:
                first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                sheets.append({"name": ws.title, "max_row": ws.max_row, "max_column": ws.max_column, "header": list(first)})
            info.update({"structural_status": "parsed_xlsx", "sheets": sheets})
        elif suffix == ".json":
            obj = json.loads(path.read_text(encoding="utf-8", errors="strict"))
            info.update({"structural_status": "parsed_json", "top_level_type": type(obj).__name__})
        elif suffix == ".zip":
            import zipfile
            with zipfile.ZipFile(path) as z:
                bad = z.testzip()
                info.update({"structural_status": "parsed_zip", "members": len(z.infolist()), "bad_member": bad})
        elif suffix in {".py", ".r", ".rmd", ".ipynb", ".md", ".txt", ".yml", ".yaml"}:
            text = path.read_text(encoding="utf-8", errors="replace")
            info.update({"structural_status": "text_readable", "line_count": text.count("\n") + 1})
    except Exception as exc:
        info.update({"structural_status": "parse_error", "parse_error": f"{type(exc).__name__}:{exc}"})
    return info


def audit_zenodo(session: requests.Session, out: Path) -> dict:
    api = "https://zenodo.org/api/records/20557057"
    r = checked_get(session, api)
    record = r.json()
    target_dir = out / "zenodo-20557057"
    target_dir.mkdir(parents=True, exist_ok=True)
    metadata = {
        "record_id": str(record.get("id")),
        "doi": record.get("doi"),
        "title": (record.get("metadata") or {}).get("title"),
        "license": ((record.get("metadata") or {}).get("license") or {}).get("id") if isinstance((record.get("metadata") or {}).get("license"), dict) else (record.get("metadata") or {}).get("license"),
        "publication_date": (record.get("metadata") or {}).get("publication_date"),
        "files": [],
        "api_url": api,
    }
    total = 0
    for item in record.get("files") or []:
        name = item.get("key") or item.get("filename")
        links = item.get("links") or {}
        url = links.get("self") or links.get("content") or links.get("download")
        entry = {"name": name, "declared_size": item.get("size"), "declared_checksum": item.get("checksum"), "url": url}
        try:
            if not name or not url:
                raise RuntimeError("missing_name_or_download_url")
            safe = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(name).name)
            path = target_dir / safe
            result = download(session, url, path)
            total += result["size_bytes"]
            if total > MAX_TOTAL_BYTES:
                raise RuntimeError("project_total_too_large")
            entry.update(result)
            entry.update(inspect_file(path))
            entry["status"] = "verified"
            md5_declared = str(item.get("checksum") or "")
            if md5_declared.startswith("md5:"):
                md5 = hashlib.md5(path.read_bytes()).hexdigest()
                entry["declared_checksum_match"] = md5 == md5_declared.split(":", 1)[1]
        except Exception as exc:
            entry.update({"status": "retryable", "error": f"{type(exc).__name__}:{exc}"})
        metadata["files"].append(entry)
    metadata["downloaded_bytes"] = total
    metadata["completed"] = bool(metadata["files"]) and all(f.get("status") == "verified" for f in metadata["files"])
    (target_dir / "audit.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    return metadata


def osf_paginate(session: requests.Session, url: str):
    while url:
        r = checked_get(session, url)
        payload = r.json()
        for item in payload.get("data") or []:
            yield item
        url = ((payload.get("links") or {}).get("next"))


def audit_osf(session: requests.Session, out: Path) -> dict:
    node_id = "tcyk9"
    base = f"https://api.osf.io/v2/nodes/{node_id}/"
    node = checked_get(session, base).json().get("data") or {}
    target_dir = out / f"osf-{node_id}"
    target_dir.mkdir(parents=True, exist_ok=True)
    attributes = node.get("attributes") or {}
    result = {
        "node_id": node_id,
        "title": attributes.get("title"),
        "public": attributes.get("public"),
        "license": attributes.get("license"),
        "date_modified": attributes.get("date_modified"),
        "api_url": base,
        "providers": [],
        "files": [],
    }
    providers_url = ((node.get("relationships") or {}).get("files") or {}).get("links", {}).get("related", {}).get("href") or f"{base}files/"
    total = 0
    for provider in osf_paginate(session, providers_url):
        p_attrs = provider.get("attributes") or {}
        p_name = p_attrs.get("name") or provider.get("id")
        result["providers"].append(p_name)
        files_url = ((provider.get("relationships") or {}).get("files") or {}).get("links", {}).get("related", {}).get("href")
        if not files_url:
            files_url = (provider.get("links") or {}).get("files")
        if not files_url:
            continue
        queue = [files_url]
        seen = set()
        while queue:
            current = queue.pop(0)
            if current in seen:
                continue
            seen.add(current)
            for item in osf_paginate(session, current):
                attrs = item.get("attributes") or {}
                kind = attrs.get("kind")
                name = attrs.get("name") or item.get("id")
                links = item.get("links") or {}
                if kind == "folder":
                    child = links.get("move") or links.get("new_folder")
                    related = ((item.get("relationships") or {}).get("files") or {}).get("links", {}).get("related", {}).get("href")
                    if related:
                        queue.append(related)
                    elif links.get("self"):
                        queue.append(links["self"])
                    continue
                url = links.get("download")
                entry = {"provider": p_name, "name": name, "kind": kind, "declared_size": attrs.get("size"), "download_url": url}
                try:
                    if not url:
                        raise RuntimeError("missing_download_url")
                    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(name).name)
                    path = target_dir / safe
                    if path.exists():
                        stem, suffix = path.stem, path.suffix
                        path = target_dir / f"{stem}_{item.get('id')}{suffix}"
                    dl = download(session, url, path)
                    total += dl["size_bytes"]
                    if total > MAX_TOTAL_BYTES:
                        raise RuntimeError("project_total_too_large")
                    entry.update(dl)
                    entry.update(inspect_file(path))
                    entry["status"] = "verified"
                except Exception as exc:
                    entry.update({"status": "retryable", "error": f"{type(exc).__name__}:{exc}"})
                result["files"].append(entry)
    result["downloaded_bytes"] = total
    result["completed"] = bool(result["files"]) and all(f.get("status") == "verified" for f in result["files"])
    (target_dir / "audit.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers.update({"User-Agent": "open-evidence-resource-audit/1.0 (lawful public research archive)"})
    summary = {"zenodo": None, "osf": None, "errors": []}
    try:
        summary["zenodo"] = audit_zenodo(session, args.output)
    except Exception as exc:
        summary["errors"].append({"source": "Zenodo", "error": f"{type(exc).__name__}:{exc}"})
    try:
        summary["osf"] = audit_osf(session, args.output)
    except Exception as exc:
        summary["errors"].append({"source": "OSF", "error": f"{type(exc).__name__}:{exc}"})
    summary["completed_sources"] = sum(bool(summary.get(k) and summary[k].get("completed")) for k in ("zenodo", "osf"))
    (args.output / "completion_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"completed_sources": summary["completed_sources"], "errors": summary["errors"]}, indent=2))
    if summary["completed_sources"] == 0:
        raise SystemExit("No resource source completed")


if __name__ == "__main__":
    main()
