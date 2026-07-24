from __future__ import annotations

import argparse
import csv
import hashlib
import json
import mimetypes
import os
import re
import shutil
import time
import zipfile
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

ZENODO_API = "https://zenodo.org/api/records/{record_id}"
MENDELEY_API = "https://api.data.mendeley.com"
MAX_DOWNLOAD_BYTES = 500 * 1024 * 1024
ALLOWED_HOSTS = {
    "zenodo.org",
    "api.data.mendeley.com",
    "prod-dcd-datasets-public-files-eu-west-1.s3.eu-west-1.amazonaws.com",
    "prod-dcd-datasets-public-files-eu-west-1.s3.amazonaws.com",
    "s3-eu-west-1.amazonaws.com",
    "amazonaws.com",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def md5_file(path: Path) -> str:
    digest = hashlib.md5()  # nosec B324 - repository checksum verification only
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_filename(name: str) -> str:
    result = Path(name).name
    if result in {"", ".", ".."}:
        raise ValueError(f"Unsafe filename: {name!r}")
    return result


def allowed_url(url: str) -> bool:
    host = (urlparse(url).hostname or "").lower()
    return host in ALLOWED_HOSTS or host.endswith(".amazonaws.com") or host.endswith(".zenodo.org")


def stream_download(session: requests.Session, url: str, output: Path) -> dict[str, Any]:
    if not allowed_url(url):
        raise RuntimeError(f"Download host is not allow-listed: {url}")
    output.parent.mkdir(parents=True, exist_ok=True)
    with session.get(url, stream=True, timeout=(30, 180), allow_redirects=True) as response:
        response.raise_for_status()
        final_url = response.url
        if not allowed_url(final_url):
            raise RuntimeError(f"Redirected to non-allow-listed host: {final_url}")
        content_length = response.headers.get("Content-Length")
        if content_length and int(content_length) > MAX_DOWNLOAD_BYTES:
            raise RuntimeError(f"File exceeds maximum size: {content_length}")
        total = 0
        with output.open("wb") as handle:
            for chunk in response.iter_content(1024 * 1024):
                if not chunk:
                    continue
                total += len(chunk)
                if total > MAX_DOWNLOAD_BYTES:
                    raise RuntimeError("File exceeded maximum size while streaming")
                handle.write(chunk)
    return {
        "requested_url": url,
        "final_url": final_url,
        "status_code": response.status_code,
        "content_type": response.headers.get("Content-Type"),
        "content_length_header": content_length,
        "size_bytes": total,
    }


def inspect_csv(path: Path) -> dict[str, Any]:
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
    last_error = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                header = next(reader)
                rows = sum(1 for _ in reader)
            return {
                "file_format": "csv",
                "encoding": encoding,
                "rows_excluding_header": rows,
                "columns": len(header),
                "column_names": header,
            }
        except (UnicodeDecodeError, StopIteration, csv.Error) as exc:
            last_error = str(exc)
    raise RuntimeError(f"CSV inspection failed: {last_error}")


def inspect_xlsx(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        try:
            first = next(iterator)
        except StopIteration:
            first = tuple()
        rows = 0
        nonempty_rows = 0
        for row in iterator:
            rows += 1
            if any(value is not None and str(value).strip() for value in row):
                nonempty_rows += 1
        headers = [None if value is None else str(value) for value in first]
        sheets.append({
            "sheet_name": sheet.title,
            "max_row_reported": sheet.max_row,
            "max_column_reported": sheet.max_column,
            "data_rows_scanned": rows,
            "nonempty_data_rows": nonempty_rows,
            "column_names": headers,
        })
    workbook.close()
    return {"file_format": "xlsx", "sheets": sheets}


def inspect_file(path: Path) -> dict[str, Any]:
    signature = path.read_bytes()[:16]
    result: dict[str, Any] = {
        "filename": path.name,
        "size_bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "md5": md5_file(path),
        "signature_hex": signature.hex(),
        "guessed_mime_type": mimetypes.guess_type(path.name)[0],
    }
    suffix = path.suffix.lower()
    if suffix == ".csv":
        result.update(inspect_csv(path))
    elif suffix == ".xlsx":
        if signature[:2] != b"PK":
            raise RuntimeError(f"XLSX is not a ZIP/OOXML file: {path}")
        result.update(inspect_xlsx(path))
    else:
        result["file_format"] = suffix.lstrip(".") or "unknown"
    return result


def match_expected(filename: str, expected: list[str]) -> bool:
    normalized = re.sub(r"\s+", " ", filename.strip().lower())
    return any(normalized == re.sub(r"\s+", " ", value.strip().lower()) for value in expected)


def download_zenodo(session: requests.Session, target: dict[str, Any], root: Path) -> dict[str, Any]:
    api_url = ZENODO_API.format(record_id=target["record_id"])
    response = session.get(api_url, timeout=(30, 120))
    response.raise_for_status()
    record = response.json()
    files = record.get("files") or []
    selected = []
    for item in files:
        key = safe_filename(item.get("key") or item.get("filename") or "")
        if not match_expected(key, target["expected_files"]):
            continue
        links = item.get("links") or {}
        url = links.get("content") or links.get("self")
        if not url:
            raise RuntimeError(f"No file content URL for {key}")
        output = root / "files" / key
        transfer = stream_download(session, url, output)
        audit = inspect_file(output)
        expected_md5 = (target.get("expected_md5") or {}).get(key)
        if expected_md5 and audit["md5"].lower() != expected_md5.lower():
            raise RuntimeError(f"MD5 mismatch for {key}")
        selected.append({"metadata": item, "transfer": transfer, "audit": audit})
    if not selected:
        raise RuntimeError(f"Expected Zenodo files not found; available: {[x.get('key') for x in files]}")
    return {
        "official_api_url": api_url,
        "record_metadata": {
            "id": record.get("id"),
            "doi": record.get("doi"),
            "title": (record.get("metadata") or {}).get("title"),
            "license": (record.get("metadata") or {}).get("license"),
            "publication_date": (record.get("metadata") or {}).get("publication_date"),
        },
        "downloaded_files": selected,
    }


def extract_expected_from_zip(zip_path: Path, expected_files: list[str], output_dir: Path) -> list[Path]:
    selected = []
    with zipfile.ZipFile(zip_path) as archive:
        members = [item for item in archive.infolist() if not item.is_dir()]
        for member in members:
            filename = safe_filename(member.filename)
            if not match_expected(filename, expected_files):
                continue
            output = output_dir / filename
            output.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(member) as source, output.open("wb") as target:
                shutil.copyfileobj(source, target)
            selected.append(output)
    return selected


def mendeley_file_objects(session: requests.Session, dataset_id: str, version: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    attempts = []
    endpoints = [
        f"{MENDELEY_API}/datasets/publics/{dataset_id}/files?version={version}&$limit=100",
        f"{MENDELEY_API}/datasets/{dataset_id}/files?version={version}&$limit=100",
    ]
    for endpoint in endpoints:
        try:
            response = session.get(endpoint, timeout=(30, 120))
            attempts.append({"url": endpoint, "status": response.status_code, "content_type": response.headers.get("Content-Type")})
            if response.status_code == 200:
                value = response.json()
                if isinstance(value, list):
                    return value, attempts
                if isinstance(value, dict):
                    for key in ("items", "data", "results"):
                        if isinstance(value.get(key), list):
                            return value[key], attempts
        except requests.RequestException as exc:
            attempts.append({"url": endpoint, "error": f"{type(exc).__name__}:{exc}"})
    return [], attempts


def download_mendeley(session: requests.Session, target: dict[str, Any], root: Path) -> dict[str, Any]:
    dataset_id = target["dataset_id"]
    version = int(target["version"])
    objects, attempts = mendeley_file_objects(session, dataset_id, version)
    selected = []
    for item in objects:
        filename = safe_filename(item.get("filename") or item.get("name") or "")
        if not match_expected(filename, target["expected_files"]):
            continue
        details = item.get("content_details") or {}
        url = details.get("download_url") or item.get("download_url")
        if not url and item.get("id"):
            details_url = f"{MENDELEY_API}/datasets/publics/{dataset_id}/files/{item['id']}?version={version}"
            details_response = session.get(details_url, timeout=(30, 120))
            attempts.append({"url": details_url, "status": details_response.status_code})
            if details_response.status_code == 200:
                detail_item = details_response.json()
                details = detail_item.get("content_details") or {}
                url = details.get("download_url") or detail_item.get("download_url")
                item = detail_item
        if not url and item.get("id"):
            url = f"{MENDELEY_API}/datasets/{dataset_id}/files/{item['id']}/file_downloaded?version={version}"
        if not url:
            continue
        output = root / "files" / filename
        transfer = stream_download(session, url, output)
        selected.append({"metadata": item, "transfer": transfer, "audit": inspect_file(output)})

    zip_audit = None
    if not selected:
        zip_url = f"{MENDELEY_API}/datasets/{dataset_id}/zip/file_downloaded?version={version}"
        zip_path = root / "downloads" / f"{dataset_id}_v{version}.zip"
        try:
            transfer = stream_download(session, zip_url, zip_path)
            if not zipfile.is_zipfile(zip_path):
                raise RuntimeError("Mendeley download-all response is not a ZIP archive")
            files = extract_expected_from_zip(zip_path, target["expected_files"], root / "files")
            zip_audit = {"transfer": transfer, "zip_sha256": sha256_file(zip_path), "zip_size": zip_path.stat().st_size}
            selected = [{"metadata": {"source": "download_all_zip"}, "audit": inspect_file(path)} for path in files]
        except Exception as exc:
            attempts.append({"url": zip_url, "error": f"{type(exc).__name__}:{exc}"})
    if not selected:
        available = [item.get("filename") or item.get("name") for item in objects]
        raise RuntimeError(f"Expected Mendeley files not downloaded. Available={available}; attempts={attempts}")
    return {
        "official_api_base": MENDELEY_API,
        "api_attempts": attempts,
        "downloaded_files": selected,
        "download_all_zip": zip_audit,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--targets", type=Path, required=True)
    parser.add_argument("--target-id", required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    targets = json.loads(args.targets.read_text(encoding="utf-8"))
    target = next((item for item in targets if item["target_id"] == args.target_id), None)
    if target is None:
        raise SystemExit(f"Unknown target id: {args.target_id}")
    root = args.output
    root.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers.update({
        "User-Agent": "ExplainabilityBiasOpenEvidence/2.0 public-data-audit",
        "Accept": "application/json, application/octet-stream, */*",
    })
    started_at = time.time()
    if target["platform"] == "Zenodo":
        source_result = download_zenodo(session, target, root)
    elif target["platform"] == "Mendeley Data":
        source_result = download_mendeley(session, target, root)
    else:
        raise SystemExit(f"Unsupported platform: {target['platform']}")
    report = {
        "target": target,
        "source_result": source_result,
        "elapsed_seconds": round(time.time() - started_at, 3),
        "completed": True,
        "integrity_boundary": "File-level availability is confirmed only after binary download, checksum and structural parsing.",
    }
    report_path = root / "audit_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    checksums = []
    for path in sorted(root.rglob("*")):
        if path.is_file() and path != report_path:
            checksums.append(f"{sha256_file(path)}  {path.relative_to(root)}\n")
    checksums.append(f"{sha256_file(report_path)}  {report_path.relative_to(root)}\n")
    (root / "sha256sums.txt").write_text("".join(checksums), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
