from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import mimetypes
import re
import time
import zipfile
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import requests
from openpyxl import load_workbook

API = "https://api.osf.io/v2"
MAX_FILE_BYTES = 250 * 1024 * 1024
MAX_TOTAL_BYTES = 800 * 1024 * 1024
DOWNLOAD_EXTENSIONS = {
    ".csv", ".tsv", ".txt", ".json", ".xlsx", ".xls", ".sav", ".dta",
    ".rds", ".rdata", ".r", ".py", ".ipynb", ".zip", ".pdf", ".docx",
}
DATA_HINT = re.compile(
    r"(data|raw|clean|response|participant|experiment|study|trial|survey|code|script|analysis|material|prereg|task)",
    re.I,
)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def add_view_only(url: str, token: str | None) -> str:
    if not token:
        return url
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["view_only"] = token
    return urlunparse(parsed._replace(query=urlencode(query)))


def save_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def get_json(session: requests.Session, url: str, view_only: str | None) -> dict[str, Any]:
    target = add_view_only(url, view_only)
    response = session.get(target, timeout=(30, 180))
    response.raise_for_status()
    value = response.json()
    if not isinstance(value, dict):
        raise RuntimeError(f"Expected JSON object from {target}")
    return value


def paginate(session: requests.Session, url: str, view_only: str | None, raw_dir: Path, label: str) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    page = 0
    next_url: str | None = url
    seen: set[str] = set()
    while next_url:
        page += 1
        target = add_view_only(next_url, view_only)
        if target in seen:
            raise RuntimeError(f"Repeated pagination URL for {label}: {target}")
        seen.add(target)
        response = session.get(target, timeout=(30, 180))
        response.raise_for_status()
        value = response.json()
        save_json(raw_dir / f"{label}_page_{page:04d}.json", value)
        data = value.get("data") or []
        if not isinstance(data, list):
            raise RuntimeError(f"Expected list data for {label}")
        output.extend(item for item in data if isinstance(item, dict))
        links = value.get("links") or {}
        next_url = links.get("next")
    return output


def safe_name(value: str) -> str:
    name = Path(value).name
    if not name or name in {".", ".."}:
        raise ValueError(value)
    return re.sub(r"[^A-Za-z0-9._()\- ]+", "_", name)


def file_attributes(item: dict[str, Any]) -> dict[str, Any]:
    attributes = item.get("attributes") or {}
    links = item.get("links") or {}
    relationships = item.get("relationships") or {}
    return {
        "id": item.get("id"),
        "name": attributes.get("name"),
        "kind": attributes.get("kind"),
        "size": attributes.get("size"),
        "materialized_path": attributes.get("materialized_path"),
        "date_created": attributes.get("date_created"),
        "date_modified": attributes.get("date_modified"),
        "provider": attributes.get("provider"),
        "path": attributes.get("path"),
        "current_version": attributes.get("current_version"),
        "extra": attributes.get("extra"),
        "download_url": links.get("download"),
        "info_url": links.get("info"),
        "files_url": (((relationships.get("files") or {}).get("links") or {}).get("related") or {}).get("href"),
    }


def recurse_files(
    session: requests.Session,
    collection_url: str,
    view_only: str | None,
    raw_dir: Path,
    label_prefix: str,
) -> list[dict[str, Any]]:
    queue: list[tuple[str, str]] = [(collection_url, label_prefix)]
    files: list[dict[str, Any]] = []
    visited: set[str] = set()
    while queue:
        url, label = queue.pop(0)
        target = add_view_only(url, view_only)
        if target in visited:
            continue
        visited.add(target)
        items = paginate(session, url, view_only, raw_dir, label)
        for item in items:
            row = file_attributes(item)
            if row["kind"] == "folder" and row.get("files_url"):
                queue.append((row["files_url"], f"{label}_folder_{len(queue)+len(visited):04d}"))
            elif row["kind"] == "file":
                files.append(row)
    return files


def should_download(row: dict[str, Any]) -> bool:
    name = str(row.get("name") or "")
    suffix = Path(name).suffix.lower()
    size = row.get("size")
    if isinstance(size, int) and size > MAX_FILE_BYTES:
        return False
    return suffix in DOWNLOAD_EXTENSIONS and (DATA_HINT.search(name) is not None or suffix in {".csv", ".tsv", ".xlsx", ".rds", ".rdata", ".sav", ".dta"})


def download(session: requests.Session, url: str, output: Path, view_only: str | None) -> dict[str, Any]:
    target = add_view_only(url, view_only)
    output.parent.mkdir(parents=True, exist_ok=True)
    total = 0
    with session.get(target, stream=True, timeout=(30, 240), allow_redirects=True) as response:
        response.raise_for_status()
        content_length = response.headers.get("Content-Length")
        if content_length and int(content_length) > MAX_FILE_BYTES:
            raise RuntimeError(f"File too large: {content_length}")
        with output.open("wb") as handle:
            for chunk in response.iter_content(1024 * 1024):
                if not chunk:
                    continue
                total += len(chunk)
                if total > MAX_FILE_BYTES:
                    raise RuntimeError("File exceeded maximum while downloading")
                handle.write(chunk)
    return {
        "requested_url": url,
        "final_url": response.url,
        "content_type": response.headers.get("Content-Type"),
        "size_bytes": total,
        "sha256": sha256_file(output),
    }


def inspect_csv(path: Path, delimiter: str = ",") -> dict[str, Any]:
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle, delimiter=delimiter)
                header = next(reader)
                rows = sum(1 for _ in reader)
            return {"format": "csv" if delimiter == "," else "tsv", "encoding": encoding, "rows": rows, "columns": len(header), "column_names": header}
        except (UnicodeDecodeError, StopIteration, csv.Error):
            continue
    return {"format": "delimited_text", "parse_status": "failed"}


def inspect_xlsx(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        try:
            header = next(iterator)
        except StopIteration:
            header = tuple()
        rows = 0
        nonempty = 0
        for row in iterator:
            rows += 1
            if any(value is not None and str(value).strip() for value in row):
                nonempty += 1
        sheets.append({"sheet_name": sheet.title, "rows": rows, "nonempty_rows": nonempty, "columns": len(header), "column_names": [None if value is None else str(value) for value in header]})
    workbook.close()
    return {"format": "xlsx", "sheets": sheets}


def inspect_file(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    signature = path.read_bytes()[:16].hex()
    base: dict[str, Any] = {"signature_hex": signature, "mime_guess": mimetypes.guess_type(path.name)[0]}
    if suffix == ".csv":
        base.update(inspect_csv(path))
    elif suffix == ".tsv":
        base.update(inspect_csv(path, "\t"))
    elif suffix == ".xlsx":
        if not path.read_bytes()[:2] == b"PK":
            raise RuntimeError(f"Invalid XLSX signature: {path}")
        base.update(inspect_xlsx(path))
    elif suffix == ".zip":
        if not zipfile.is_zipfile(path):
            raise RuntimeError(f"Invalid ZIP: {path}")
        with zipfile.ZipFile(path) as archive:
            members = [{"name": item.filename, "size": item.file_size, "compressed_size": item.compress_size} for item in archive.infolist() if not item.is_dir()]
        base.update({"format": "zip", "members": members})
    else:
        base.update({"format": suffix.lstrip(".") or "unknown", "structural_parse": "not_attempted_no_execution"})
    return base


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--targets", type=Path, required=True)
    parser.add_argument("--target-id", required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    targets = json.loads(args.targets.read_text(encoding="utf-8"))
    target = next((item for item in targets if item["target_id"] == args.target_id), None)
    if target is None:
        raise SystemExit(f"Unknown target {args.target_id}")
    root = args.output
    raw_dir = root / "raw_api"
    files_dir = root / "files"
    root.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers.update({"User-Agent": "ExplainabilityBiasOpenEvidence/2.0 OSF-public-file-audit", "Accept": "application/vnd.api+json, application/json, */*"})
    view_only = target.get("view_only")
    node_url = f"{API}/nodes/{target['node_id']}/"
    node = get_json(session, node_url, view_only)
    save_json(raw_dir / "node.json", node)
    node_data = node.get("data") or {}
    node_attributes = node_data.get("attributes") or {}
    files_relationship = (((node_data.get("relationships") or {}).get("files") or {}).get("links") or {}).get("related") or {}
    providers_url = files_relationship.get("href")
    if not providers_url:
        raise RuntimeError("OSF node did not expose a files relationship")
    providers = paginate(session, providers_url, view_only, raw_dir, "providers")
    all_files: list[dict[str, Any]] = []
    for index, provider in enumerate(providers, start=1):
        provider_row = file_attributes(provider)
        provider_files_url = provider_row.get("files_url") or (provider.get("links") or {}).get("new_folder")
        if provider_files_url:
            all_files.extend(recurse_files(session, provider_files_url, view_only, raw_dir, f"provider_{index:02d}"))

    downloaded = []
    errors = []
    total_bytes = 0
    for index, row in enumerate(all_files, start=1):
        if not should_download(row) or not row.get("download_url"):
            continue
        size = row.get("size") if isinstance(row.get("size"), int) else 0
        if total_bytes + size > MAX_TOTAL_BYTES:
            errors.append({"name": row.get("name"), "error": "project_download_budget_exceeded"})
            continue
        filename = f"{index:04d}_{safe_name(str(row.get('name') or row.get('id')))}"
        output = files_dir / filename
        try:
            transfer = download(session, row["download_url"], output, view_only)
            audit = inspect_file(output)
        except Exception as exc:
            errors.append({"name": row.get("name"), "error": f"{type(exc).__name__}:{exc}"})
            continue
        total_bytes += transfer["size_bytes"]
        downloaded.append({"osf_file": row, "local_file": str(output.relative_to(root)), "transfer": transfer, "audit": audit})

    report = {
        "target": target,
        "node": {
            "id": node_data.get("id"),
            "title": node_attributes.get("title"),
            "public": node_attributes.get("public"),
            "date_created": node_attributes.get("date_created"),
            "date_modified": node_attributes.get("date_modified"),
            "category": node_attributes.get("category"),
            "registration": node_attributes.get("registration"),
        },
        "providers": len(providers),
        "files_discovered": len(all_files),
        "files_selected_and_downloaded": len(downloaded),
        "downloaded_bytes": total_bytes,
        "downloaded": downloaded,
        "errors": errors,
        "completed": True,
        "participant_level_verified": False,
        "trial_level_verified": False,
        "eligibility_note": "Participant/trial availability remains false until column-level audit demonstrates real participant or trial records.",
        "official_api": API,
        "view_only_token_used": bool(view_only),
        "generated_at": time.time(),
    }
    save_json(root / "audit_report.json", report)
    checksums = []
    for path in sorted(root.rglob("*")):
        if path.is_file() and path.name != "sha256sums.txt":
            checksums.append(f"{sha256_file(path)}  {path.relative_to(root)}\n")
    (root / "sha256sums.txt").write_text("".join(checksums), encoding="utf-8")
    print(json.dumps({"target_id": args.target_id, "files_discovered": len(all_files), "downloaded": len(downloaded), "errors": len(errors), "completed": True}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
