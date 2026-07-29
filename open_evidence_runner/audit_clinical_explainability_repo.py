from __future__ import annotations

import argparse
import hashlib
import json
import re
import tarfile
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook

REPOSITORY = "junnsang/Explainability"
COMMIT = "1c48b4f6ce1f612833ce448537bdf02cdfcc758f"
ARCHIVE_URL = f"https://api.github.com/repos/{REPOSITORY}/tarball/{COMMIT}"
MAX_BYTES = 100 * 1024 * 1024
SENSITIVE = re.compile(r"(^|_)(email|name|phone|address|ip|workerid|participantid|patientid|mrn)($|_)", re.I)
BEHAVIOR = re.compile(r"(participant|subject|doctor|clinician|accept|trust|satisf|usab|decision|choice|condition|explanation|shap|advice|vignette|response)", re.I)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def inspect_csv(path: Path) -> dict:
    frame = pd.read_csv(path, low_memory=False)
    cols = [str(c) for c in frame.columns]
    return {
        "rows": int(len(frame)),
        "columns": len(cols),
        "column_names": cols,
        "sensitive_column_candidates": [c for c in cols if SENSITIVE.search(c)],
        "behavioral_column_candidates": [c for c in cols if BEHAVIOR.search(c)],
        "duplicate_rows": int(frame.astype(str).duplicated().sum()),
        "null_counts": {str(k): int(v) for k, v in frame.isna().sum().items()},
    }


def inspect_xlsx(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for ws in workbook.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            header = tuple()
        rows = sum(1 for _ in it)
        cols = ["" if value is None else str(value) for value in header]
        sheets.append({
            "sheet": ws.title,
            "rows": rows,
            "columns": len(cols),
            "column_names": cols,
            "sensitive_column_candidates": [c for c in cols if SENSITIVE.search(c)],
            "behavioral_column_candidates": [c for c in cols if BEHAVIOR.search(c)],
        })
    workbook.close()
    return {"sheets": sheets}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)
    archive = args.output / f"Explainability-{COMMIT}.tar.gz"
    response = requests.get(ARCHIVE_URL, headers={"Accept": "application/vnd.github+json", "User-Agent": "open-evidence-clinical-repo-audit/1.0"}, stream=True, allow_redirects=True, timeout=(30, 240))
    response.raise_for_status()
    total = 0
    with archive.open("wb") as fh:
        for chunk in response.iter_content(1024 * 1024):
            if not chunk:
                continue
            total += len(chunk)
            if total > MAX_BYTES:
                raise RuntimeError("archive_exceeds_limit")
            fh.write(chunk)
    if archive.read_bytes()[:2] != b"\x1f\x8b":
        raise RuntimeError("not_gzip_archive")

    extract_root = args.output / "snapshot"
    manifest = []
    audits = []
    license_text = None
    readme_text = None
    code_files = []
    secret_candidates = []
    secret_pattern = re.compile(r"(?i)(api[_-]?key|secret|token|password)\s*[=:]\s*['\"][^'\"]{8,}")
    with tarfile.open(archive, "r:gz") as tf:
        for member in tf.getmembers():
            if not member.isfile():
                continue
            relative = "/".join(member.name.split("/")[1:])
            manifest.append({"path": relative, "size": member.size})
            extracted = tf.extractfile(member)
            raw = extracted.read() if extracted else b""
            lower = relative.lower()
            if Path(lower).name in {"license", "license.txt", "license.md"}:
                license_text = raw.decode("utf-8", errors="replace")
            if Path(lower).name == "readme.md":
                readme_text = raw.decode("utf-8", errors="replace")
            if lower.endswith((".py", ".r", ".rmd", ".ipynb")):
                code_files.append(relative)
                if secret_pattern.search(raw.decode("utf-8", errors="ignore")):
                    secret_candidates.append(relative)
            if lower.endswith((".csv", ".xlsx")):
                path = extract_root / relative
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_bytes(raw)
                try:
                    audit = inspect_csv(path) if lower.endswith(".csv") else inspect_xlsx(path)
                    audits.append({"path": relative, "size": member.size, "sha256": sha256(path), "audit": audit})
                except Exception as exc:
                    audits.append({"path": relative, "size": member.size, "sha256": sha256(path), "error": f"{type(exc).__name__}:{exc}"})

    participant_candidates = []
    sensitive_candidates = []
    for item in audits:
        audit = item.get("audit") or {}
        if audit.get("behavioral_column_candidates"):
            participant_candidates.append(item["path"])
        for sheet in audit.get("sheets", []):
            if sheet.get("behavioral_column_candidates"):
                participant_candidates.append(item["path"])
            if sheet.get("sensitive_column_candidates"):
                sensitive_candidates.append(item["path"])
        if audit.get("sensitive_column_candidates"):
            sensitive_candidates.append(item["path"])

    summary = {
        "repository": REPOSITORY,
        "fixed_commit": COMMIT,
        "archive_url": ARCHIVE_URL,
        "final_url": response.url,
        "archive_bytes": archive.stat().st_size,
        "archive_sha256": sha256(archive),
        "file_count": len(manifest),
        "data_file_count": len(audits),
        "code_files": code_files,
        "license_status": "present" if license_text else "not_declared",
        "readme_present": readme_text is not None,
        "participant_level_candidate_files": sorted(set(participant_candidates)),
        "sensitive_column_candidate_files": sorted(set(sensitive_candidates)),
        "embedded_secret_candidates": secret_candidates,
        "data_audits": audits,
        "code_executed": False,
    }
    (args.output / "repository_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    (args.output / "completion_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (args.output / "SHA256SUMS.txt").write_text(f"{summary['archive_sha256']}  {archive.name}\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if secret_candidates:
        raise SystemExit("Potential embedded secret requires manual review")
    if not audits:
        raise SystemExit("No data file was found")


if __name__ == "__main__":
    main()
